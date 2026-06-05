import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\Administrator\Desktop\MABINOGI-COLOR-SEARCH-main")
SOURCE = ROOT / "指定染色劑(ver.251112) 的副本.xlsx"
OUTPUT = ROOT / "dye-series-data.js"

HEX_RE = re.compile(r"^#?[0-9A-Fa-f]{6}$")


def normalize_hex(value):
    if value is None:
        return None
    text = str(value).strip()
    if not HEX_RE.match(text):
        return None
    return text.upper().replace("#", "")


def add_entry(index, raw_hex, series, source, group=None, note=None):
    hex_code = normalize_hex(raw_hex)
    if not hex_code or not series:
        return

    entry = {
        "series": str(series).strip(),
        "source": source,
    }
    if group:
        entry["group"] = group
    if note:
        entry["note"] = note

    bucket = index.setdefault(hex_code, [])
    if entry not in bucket:
        bucket.append(entry)


def parse_clothing_sheet(ws, index):
    current_name = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_name = row[2]
        if row_name:
            current_name = str(row_name).strip()
        if not current_name:
            continue

        for value in row[4:9]:
            add_entry(index, value, current_name, "服裝染色劑組合", "指定染")
        for value in row[10:15]:
            add_entry(index, value, current_name, "服裝染色劑組合", "指定金屬染")


def parse_single_color_sheet(ws, index):
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        series = row[2]
        if not series:
            continue
        for col_idx, value in enumerate(row[3:], start=3):
            add_entry(index, value, str(series).strip(), "單色組合", headers[col_idx])


def parse_palette_sheet(ws, index):
    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for start in (1, 4, 7):
            series = row[start]
            color = row[start + 1]
            chance = row[start - 1] if start - 1 >= 0 else None
            note = f"機率 {chance}" if chance not in (None, "") else None
            add_entry(index, color, series, "時尚彩繪調色盤", headers[start], note)


def parse_hair_sheet(ws, index):
    # row 2 stores group names, every block starts at columns 1,5,9...
    for start_col in range(1, ws.max_column + 1, 4):
        group_name = ws.cell(row=2, column=start_col).value
        if not group_name:
            continue
        group_name = str(group_name).strip()
        for row_idx in range(4, ws.max_row + 1):
            add_entry(
                index,
                ws.cell(row=row_idx, column=start_col).value,
                group_name,
                "髮染箱&轉蛋髮染",
            )


def parse_sheet_as_named_columns(ws, index, source, column_pairs):
    for row in ws.iter_rows(min_row=2, values_only=True):
        for series_name, color_idx, group in column_pairs:
            add_entry(index, row[color_idx], series_name, source, group)


def build_index():
    wb = load_workbook(SOURCE, data_only=True)
    index = {}

    parse_clothing_sheet(wb["服裝染色劑組合"], index)
    parse_single_color_sheet(wb["單色組合"], index)
    parse_palette_sheet(wb["時尚彩繪調色盤"], index)
    parse_hair_sheet(wb["髮染箱&轉蛋髮染"], index)

    parse_sheet_as_named_columns(
        wb["로나의 염색 앰플 세트 상자(寵染)"],
        index,
        "로나의 염색 앰플 세트 상자(寵染)",
        [
            ("로나의 指定染(不可交易)", 0, "指定染"),
            ("로나의 指定金屬染(不可交易)", 5, "指定金屬染"),
        ],
    )

    parse_sheet_as_named_columns(
        wb["染(寵, 樂器, 名字)"],
        index,
        "染色分類",
        [
            ("寵物染色", 0, "寵染"),
            ("樂器染色", 5, "樂器染"),
            ("名字聊天染(30天)", 11, "名字聊天染"),
        ],
    )

    return {
        "version": "251112",
        "sourceFile": SOURCE.name,
        "colorCount": len(index),
        "seriesByHex": dict(sorted(index.items())),
    }


def main():
    data = build_index()
    payload = "window.DYE_SERIES_DATA = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";\n"
    OUTPUT.write_text(payload, encoding="utf-8")
    print(f"Wrote {OUTPUT} with {data['colorCount']} colors")


if __name__ == "__main__":
    main()
